import json
import logging
from ortools.linear_solver import pywraplp
from openpyxl import Workbook

# Logging Configuration
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("debug.log"), logging.StreamHandler()],
)

class LayoutOptimizer:
    def __init__(self, input_data):
        self.config = input_data['general_configuration']
        self.layouts = input_data['layouts']
        self.fabrics = {f['fabric']: f for f in input_data['fabrics']}
        self.orders = input_data['production_orders']
        self.sizes = ["P", "M", "G", "GG"]
        self.overproduction_penalty = self.config.get('overproduction_percentage', 0.05)
        self.unit_waste_cost = self.config.get('waste_cost', 0.1)  # Ensure 'waste_cost' exists in JSON

    def calculate_layout_costs(self, layout, num_layers, used):
        fabric = self.fabrics[layout['fabric']]
        
        # Fixed Costs (applied if the layout is used)
        setup_cost = fabric['cost_per_layer'] * num_layers
        
        # Variable Costs (depend on whether the layout is used)
        length_meters = (layout['layout_length'] / 1000)
        perimeter_meters = (layout['total_perimeter'] / 1000)
        waste_area_m2 = (layout['waste_area'] / 1_000_000)
        
        fabric_cost = length_meters * fabric['price_per_linear_meter'] * used
        cutting_cost = perimeter_meters * fabric['cost_per_cut_meter'] * used
        layout_cost = length_meters * fabric['cost_per_layout_meter'] * used
        
        # Waste Cost
        waste_cost = waste_area_m2 * self.unit_waste_cost * used
        
        # Total Cost
        total_cost = setup_cost + fabric_cost + cutting_cost + layout_cost + waste_cost
        
        # Return all components
        return {
            'total_cost': total_cost,
            'setup_cost': setup_cost,
            'fabric_cost': fabric_cost,
            'cutting_cost': cutting_cost,
            'layout_cost': layout_cost,
            'waste_cost': waste_cost,
            'fabric_meters': length_meters * used,
            'cut_perimeter': perimeter_meters * used,
            'fabric_waste': waste_area_m2 * used  # Changed key to 'fabric_waste'
        }

    def preprocess_layouts(self, demand, fabric_width):
        """Preprocess and filter layouts compatible with the demand and fabric width"""
        filtered_layouts = []
        for layout in self.layouts:
            # Filter by fabric width
            if layout['fabric_width'] != fabric_width:
                continue
            
            # Check if the layout contains the necessary patterns
            layout_patterns = [p['pattern'] for p in layout['pieces']]
            order_patterns = [p['pattern'] for p in demand['pieces']]
            if not set(order_patterns).issubset(set(layout_patterns)):
                continue
            
            # Calculate efficiency per size
            layout['efficiency'] = {}
            for size in self.sizes:
                total_pieces_size = sum(
                    p['size_grade'].get(size, 0) for p in layout['pieces']
                )
                pieces_per_meter = total_pieces_size / (layout['layout_length'] / 1000)
                efficiency = pieces_per_meter * layout['utilization']
                layout['efficiency'][size] = efficiency

            filtered_layouts.append(layout)
                
        return filtered_layouts

    def optimize_order(self, order):
        demand = order['demand'][0]
        order_pieces = demand['pieces']
        demand_quantity = {}
        for piece in order_pieces:
            pattern = piece['pattern']
            for size in self.sizes:
                demand_quantity[size] = demand_quantity.get(size, 0) + piece['quantity'].get(size, 0)

        max_order_layers = demand.get('max_layers', self.config['max_layers'])
        max_order_length = demand.get('max_length', self.config['max_total_length'])
        fabric_width = demand['fabric_width']
        
        # Preprocessing
        filtered_layouts = self.preprocess_layouts(demand, fabric_width)
        if not filtered_layouts:
            logging.warning(f"No compatible layouts found for order {order['id']}")
            return None

        # Create solver
        solver = pywraplp.Solver.CreateSolver("SCIP")
        if not solver:
            return None
                
        solver.SetTimeLimit(600_000)  # 10 minutes
        
        # Decision Variables
        x = {}  # number of layers per layout
        y = {}  # binary variable indicating if the layout is used
        for layout in filtered_layouts:
            max_layout_layers = min(max_order_layers, self.config['max_layers'])
            x[layout['id']] = solver.IntVar(0, max_layout_layers, f'x_{layout["id"]}')
            y[layout['id']] = solver.IntVar(0, 1, f'y_{layout["id"]}')
            # Link x and y
            solver.Add(x[layout['id']] <= max_layout_layers * y[layout['id']], f'link_{layout["id"]}')
                
        # Demand Constraints by Size
        production = {}
        overproduction = {}
        for size in self.sizes:
            production[size] = solver.Sum([
                x[layout['id']] * sum(
                    p['size_grade'].get(size, 0) for p in layout['pieces'] if p['pattern'] in [piece['pattern'] for piece in order_pieces]
                )
                for layout in filtered_layouts
            ])
                
            # Minimum Demand
            solver.Add(production[size] >= demand_quantity.get(size, 0), f'min_demand_{size}')
            
            # Overproduction
            excess = solver.NumVar(0, solver.infinity(), f'excess_{size}')
            overproduction[size] = excess
            solver.Add(production[size] - demand_quantity.get(size, 0) == excess, f'excess_def_{size}')
            # Overproduction Limit
            max_overproduction = demand_quantity.get(size, 0) * self.config['overproduction_percentage']
            solver.Add(excess <= max_overproduction, f'max_overproduction_{size}')
                
        # Total Length Constraint
        total_length = solver.Sum([
            y[layout['id']] * layout['layout_length']
            for layout in filtered_layouts
        ])
        solver.Add(total_length <= max_order_length, 'max_length')
            
        # Objective Function
        total_cost = solver.Sum([
            self.calculate_layout_costs(layout, x[layout['id']], y[layout['id']])['total_cost']
            for layout in filtered_layouts
        ])
        
        overproduction_penalty = solver.Sum([
            overproduction[size] * self.overproduction_penalty
            for size in self.sizes
        ])
        
        # Objective: minimize total cost + penalties
        solver.Minimize(total_cost + overproduction_penalty)
            
        # Solve
        status = solver.Solve()
            
        if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
            solution = {
                layout['id']: int(x[layout['id']].solution_value())
                for layout in filtered_layouts
                if y[layout['id']].solution_value() > 0
            }
                
            # Calculate Metrics
            actual_production = {size: 0 for size in self.sizes}
            total_cost_val = 0
            total_waste_area = 0
            total_length_val = 0
            total_cutting_cost = 0
            total_layout_cost = 0
            total_waste_cost = 0
            total_fabric_meters = 0
            total_cut_perimeter = 0
                
            for layout_id, layers in solution.items():
                layout = next(l for l in filtered_layouts if l['id'] == layout_id)
                    
                for size in self.sizes:
                    total_pieces_size = sum(
                        p['size_grade'].get(size, 0) for p in layout['pieces'] if p['pattern'] in [piece['pattern'] for piece in order_pieces]
                    )
                    actual_production[size] += total_pieces_size * layers
                        
                used = y[layout_id].solution_value()
                costs = self.calculate_layout_costs(layout, layers, used)
                total_cost_val += costs['total_cost']
                total_waste_area += costs['fabric_waste'] * 1_000_000  # Converting from m² to mm²
                total_length_val += costs['fabric_meters'] * 1000  # Converting to mm
                total_cutting_cost += costs['cutting_cost']
                total_layout_cost += costs['layout_cost']
                total_waste_cost += costs['waste_cost']
                total_fabric_meters += costs['fabric_meters']
                total_cut_perimeter += costs['cut_perimeter']
                
            # Calculate Overproduction
            actual_overproduction = {size: max(0, actual_production[size] - demand_quantity.get(size, 0)) for size in self.sizes}
                
            # Collect Data for JSON
            production_pieces = []
            for piece in order_pieces:
                pattern = piece['pattern']
                fabrics = piece['fabrics']
                quantity = piece['quantity']
                production_piece = {size: 0 for size in self.sizes}
                used_layout_ids = []

                for layout_id, layers in solution.items():
                    layout = next(l for l in filtered_layouts if l['id'] == layout_id)
                    for p in layout['pieces']:
                        if p['pattern'] == pattern:
                            for size in self.sizes:
                                qty = p['size_grade'].get(size, 0) * layers
                                production_piece[size] += qty
                            if layout_id not in used_layout_ids:
                                used_layout_ids.append(layout_id)
                production_pieces.append({
                    "pattern": pattern,
                    "fabrics": fabrics,
                    "quantity": quantity,
                    "production": production_piece,
                    "layout_ids": used_layout_ids
                })

            # Collect Used Layouts
            used_layouts = []
            for layout_id, layers in solution.items():
                layout = next(l for l in self.layouts if l['id'] == layout_id)
                layout_copy = layout.copy()
                layout_copy['layers'] = layers
                layout_copy['order_ids'] = [order['id']]
                used_layouts.append(layout_copy)

            # KPIs
            kpis = {
                "total_cost": total_cost_val,
                "cutting_cost": total_cutting_cost,
                "layout_cost": total_layout_cost,
                "waste_cost": total_waste_cost,
                "fabric_meters": total_fabric_meters,
                "cut_perimeter": total_cut_perimeter,
                "fabric_waste": total_waste_area / 1_000_000  # Converting to m²
            }

            return {
                "status": "optimal",
                "solution": solution,
                "production": actual_production,
                "overproduction": actual_overproduction,
                "metrics": {
                    "total_cost": total_cost_val,
                    "cutting_cost": total_cutting_cost,
                    "layout_cost": total_layout_cost,
                    "waste_cost": total_waste_cost,
                    "fabric_meters": total_fabric_meters,
                    "cut_perimeter": total_cut_perimeter,
                    "fabric_waste": total_waste_area,
                    "total_length": total_length_val,
                    "overproduction_penalty": sum(
                        actual_overproduction[size] * self.overproduction_penalty for size in self.sizes
                    )
                },
                "kpis": kpis,
                "production_pieces": production_pieces,
                "used_layouts": used_layouts
            }
                
        return None

    def export_results(self, results):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
            
        # Headers
        headers = [
            "Order", "Layout", "Layers", "P", "M", "G", "GG",
            "Total Length (m)", "Cut Perimeter (m)", "Waste Area (m²)",
            "Cutting Cost (R$)", "Layout Cost (R$)", "Waste Cost (R$)", "Total Cost (R$)"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
            
        row = 2
        for order_id, result in results.items():
            if result:
                for layout_id, layers in result['solution'].items():
                    layout = next(l for l in self.layouts if l['id'] == layout_id)
                        
                    used = 1  # Layout was used
                    costs = self.calculate_layout_costs(layout, layers, used)
                    
                    ws.cell(row=row, column=1, value=order_id)
                    ws.cell(row=row, column=2, value=layout_id)
                    ws.cell(row=row, column=3, value=layers)
                        
                    for i, size in enumerate(self.sizes, 4):
                        piece_quantity = sum(
                            p['size_grade'].get(size, 0) for p in layout['pieces']
                            if p['pattern'] in [piece['pattern'] for piece in self.orders[int(order_id)-1]['demand'][0]['pieces']]
                        ) * layers
                        ws.cell(row=row, column=i, value=piece_quantity)
                        
                    length_meters = costs['fabric_meters']
                    perimeter_meters = costs['cut_perimeter']
                    fabric_waste_m2 = costs['fabric_waste']  # Updated key
                    
                    ws.cell(row=row, column=8, value=round(length_meters, 2))
                    ws.cell(row=row, column=9, value=round(perimeter_meters, 2))
                    ws.cell(row=row, column=10, value=round(fabric_waste_m2, 4))  # Updated key
                    
                    ws.cell(row=row, column=11, value=round(costs['cutting_cost'], 2))
                    ws.cell(row=row, column=12, value=round(costs['layout_cost'], 2))
                    ws.cell(row=row, column=13, value=round(costs['waste_cost'], 2))
                    ws.cell(row=row, column=14, value=round(costs['total_cost'], 2))
                        
                    row += 1
                
                # Add overall order metrics
                row += 1
                ws.cell(row=row, column=1, value="Order Totals")
                ws.cell(row=row, column=3, value="Total")
                for i, size in enumerate(self.sizes, 4):
                    ws.cell(row=row, column=i, value=result['production'][size])
                ws.cell(row=row, column=8, value=round(result['metrics']['fabric_meters'], 2))
                ws.cell(row=row, column=9, value=round(result['metrics']['cut_perimeter'], 2))
                ws.cell(row=row, column=10, value=round(result['metrics']['fabric_waste'], 4))  # Updated key
                ws.cell(row=row, column=11, value=round(result['metrics']['cutting_cost'], 2))
                ws.cell(row=row, column=12, value=round(result['metrics']['layout_cost'], 2))
                ws.cell(row=row, column=13, value=round(result['metrics']['waste_cost'], 2))
                ws.cell(row=row, column=14, value=round(result['metrics']['total_cost'], 2))
                row += 2  # Space between orders
            else:
                # If no solution, log in Excel
                ws.cell(row=row, column=1, value=order_id)
                ws.cell(row=row, column=2, value="No solution")
                row += 1
        
        wb.save("results.xlsx")

    def export_results_json(self, results):
        results_json = []
        for order_id, result in results.items():
            if result:
                order_json = {
                    "kpis": result['kpis'],
                    "production": result['production_pieces'],
                    "layouts": result['used_layouts']
                }
                results_json.append(order_json)
        with open('results.json', 'w', encoding='utf-8') as f:
            json.dump(results_json, f, ensure_ascii=False, indent=4)

def main():
    # Load data
    with open('dados_entrada.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Create optimizer
    optimizer = LayoutOptimizer(data)
    
    # Process orders
    results = {}
    for order in optimizer.orders:
        logging.info(f"Optimizing order {order['id']}")
        result = optimizer.optimize_order(order)
        results[order['id']] = result
        
        if result:
            logging.info(f"Solution found for order {order['id']}")
            # Production details...
            for size in optimizer.sizes:
                logging.info(f"  {size}: {result['production'][size]} pieces (Excess: {result['overproduction'][size]})")
            logging.info(f"Total cost: R$ {result['metrics']['total_cost']:.2f}")
            logging.info(f"Cutting cost: R$ {result['metrics']['cutting_cost']:.2f}")
            logging.info(f"Layout cost: R$ {result['metrics']['layout_cost']:.2f}")
            logging.info(f"Waste cost: R$ {result['metrics']['waste_cost']:.2f}")
            logging.info(f"Fabric meters: {result['metrics']['fabric_meters']:.2f} m")
            logging.info(f"Cut perimeter: {result['metrics']['cut_perimeter']:.2f} m")
            logging.info(f"Fabric waste: {result['metrics']['fabric_waste'] / 1_000_000:.4f} m²")  # Updated key
            logging.info(f"Overproduction penalty: R$ {result['metrics']['overproduction_penalty']:.2f}")
        else:
            logging.warning(f"Could not find a solution for order {order['id']}")
    
    # Export results
    if any(results.values()):
        optimizer.export_results(results)
        optimizer.export_results_json(results)
        logging.info("Results exported to results.xlsx and results.json")
    else:
        logging.warning("No solutions found for any orders")

if __name__ == "__main__":
    main()
