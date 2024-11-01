import logging
from ortools.linear_solver import pywraplp
import random
import datetime
import plotly.figure_factory as ff
import numpy as np


from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import plotly.io as pio
import os


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[logging.FileHandler("debug.log"), logging.StreamHandler()],
)


class Recurso:
    def __init__(self, id, eficiencia):
        self.id = id
        self.eficiencia = eficiencia


class Turno:
    def __init__(self, inicio, fim, eficiencia=1.0):
        self.inicio = inicio
        self.fim = fim
        self.eficiencia = eficiencia


def salvar_figura_html(fig, caminho_arquivo="gantt_chart.html"):
    try:
        fig.write_html(caminho_arquivo)
        logging.info(f"Figura HTML salva em {caminho_arquivo}")
    except Exception as e:
        logging.error(f"Erro ao salvar a figura HTML: {e}")


def exportar_para_excel(
    cronograma,
    pedidos_ordenados,
    resultados,
    pedidos,
    tamanhos,
    grades,
    fig,
    nome_arquivo=r"G:\Meu Drive\senai_sc\resultados_producao.xlsx",
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cronograma de Produção"

    # Estilo para cabeçalhos
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
    )

    # Cronograma
    ws.cell(row=1, column=1, value="Cronograma de Produção").font = Font(
        bold=True, size=14
    )
    headers = ["Pedido", "Operação", "Início", "Fim", "Recurso"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    row = 4
    for p in pedidos_ordenados:
        ws.cell(row=row, column=1, value=f"Pedido {p}")
        ws.cell(row=row, column=2, value="Enfestamento")
        ws.cell(row=row, column=3, value=cronograma[p]["inicio_enfestamento"])
        ws.cell(row=row, column=4, value=cronograma[p]["fim_enfestamento"])
        ws.cell(
            row=row, column=5, value=f"Enfestadeira {cronograma[p]['enfestadeira']}"
        )
        row += 1

        ws.cell(row=row, column=1, value=f"Pedido {p}")
        ws.cell(row=row, column=2, value="Corte")
        ws.cell(row=row, column=3, value=cronograma[p]["inicio_corte"])
        ws.cell(row=row, column=4, value=cronograma[p]["fim_corte"])
        ws.cell(
            row=row,
            column=5,
            value=f"Máquina de Corte {cronograma[p]['maquina_corte']}",
        )
        row += 1

    # Detalhes dos pedidos
    row += 2
    ws.cell(row=row, column=1, value="Detalhes dos Pedidos").font = Font(
        bold=True, size=14
    )
    row += 1
    headers = [
        "Pedido",
        "Prazo",
        "Custo Total",
        "Custo de Setup",
        "Metros de Tecido",
        "Perímetro Cortado",
        "Desperdício",
        "Comprimento Enfesto real",
        "Comprimento Enfesto Máximo",
        "Relaxação",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    for p in pedidos_ordenados:
        resultado = resultados[p]
        ws.cell(row=row, column=1, value=f"Pedido {p}")
        ws.cell(row=row, column=2, value=pedidos[p]["prazo"])
        ws.cell(row=row, column=3, value=resultado["custo_total"])
        ws.cell(row=row, column=4, value=resultado["custo_setup"])
        ws.cell(row=row, column=5, value=resultado["metros_tecido"])
        ws.cell(row=row, column=6, value=resultado["perimetro_cortado"])
        ws.cell(row=row, column=7, value=resultado["desperdicio"])
        # ws.cell(row=row, column=8, value=resultado['comprimento_enfesto_real'])
        ws.cell(row=row, column=8, value=resultado["comprimento_enfesto_maximo"])
        ws.cell(row=row, column=9, value=resultado["relaxacao_aplicada"])
        row += 1

    # Produção por tamanho
    row += 2
    ws.cell(row=row, column=1, value="Produção por Tamanho").font = Font(
        bold=True, size=14
    )
    row += 1
    headers = ["Pedido"] + tamanhos + ["Grade", "Camadas"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    row += 1

    for p in pedidos_ordenados:
        resultado = resultados[p]
        ws.cell(row=row, column=1, value=f"Pedido {p}")
        for col, t in enumerate(tamanhos, start=2):
            ws.cell(
                row=row,
                column=col,
                value=f"{int(resultado['producao'][t])} ({pedidos[p]['demandas'][t]})",
            )

        # Adicionar informações das grades e camadas
        for g, camadas in resultado["camadas"].items():
            if camadas > 0:
                ws.cell(row=row, column=len(tamanhos) + 2, value=g)
                ws.cell(row=row, column=len(tamanhos) + 3, value=int(camadas))
                row += 1

        row += 1

    # Ajustar largura das colunas
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Exportar e inserir o gráfico de Gantt
    # pio.write_image(fig, "gantt_chart.png", scale=2)
    # pio.write_image(fig, "gantt_chart.svg", format="svg")
    # img = Image("gantt_chart.svg")
    # ws.add_image(img, f"A{row + 2}")

    wb.save(nome_arquivo)
    print(f"Resultados exportados para {nome_arquivo}")
    # os.remove("gantt_chart.svg")


def otimizar_pedido_com_relaxacao(
    pedido,
    grades,
    tamanhos,
    comprimento_mesa_enfesto,
    recursos,
    percentual_superproducao,
    max_camadas_por_grade,
    horas_producao,
    penalizacao_superproducao,
    relaxacao,
):

    if relaxacao == True:

        num_pontos = int((2 - 1) / 0.1) + 1
        print(num_pontos)

        relaxacao_list = np.linspace(2, 1, num=num_pontos)[::-1].tolist()
        print(relaxacao_list)
    else:
        relaxacao_list = [1.0]

    for relaxacao in relaxacao_list:
        resultado = otimizar_pedido(
            pedido,
            grades,
            tamanhos,
            comprimento_mesa_enfesto,
            recursos,
            percentual_superproducao,
            max_camadas_por_grade,
            fator_relaxacao=relaxacao,
            horas_producao=horas_producao,
            penalizacao_superproducao=penalizacao_superproducao,
        )
        if resultado:
            resultado["relaxacao_aplicada"] = relaxacao
            return resultado
    return None


def criar_grafico_gantt(cronograma, pedidos_ordenados, criterio_prioridade):
    df = []
    for p in pedidos_ordenados:
        df.append(
            dict(
                Task=f"Pedido {p} - Enfestamento",
                Start=cronograma[p]["inicio_enfestamento"],
                Finish=cronograma[p]["fim_enfestamento"],
                Resource=f"Enfestadeira {cronograma[p]['enfestadeira']}",
            )
        )
        df.append(
            dict(
                Task=f"Pedido {p} - Corte",
                Start=cronograma[p]["inicio_corte"],
                Finish=cronograma[p]["fim_corte"],
                Resource=f"Máquina de Corte {cronograma[p]['maquina_corte']}",
            )
        )

    fig = ff.create_gantt(
        df, index_col="Resource", show_colorbar=True, group_tasks=True
    )
    fig.update_layout(
        title=f"Cronograma de Produção - {criterio_prioridade}",
        xaxis_title="Data",
        yaxis_title="Recursos e Pedidos",
    )
    return fig


def calcular_tempo_producao(resultado, grades, recursos):
    tempo_enfesto = resultado["tempo_enfesto"]
    tempo_corte = resultado["tempo_corte"]

    eficiencia_media_enfestadeiras = sum(
        r.eficiencia for r in recursos["enfestadeiras"]
    ) / len(recursos["enfestadeiras"])
    eficiencia_media_maquinas_corte = sum(
        r.eficiencia for r in recursos["maquinas_corte"]
    ) / len(recursos["maquinas_corte"])

    tempo_enfesto_ajustado = tempo_enfesto / (
        len(recursos["enfestadeiras"]) * eficiencia_media_enfestadeiras
    )
    tempo_corte_ajustado = tempo_corte / (
        len(recursos["maquinas_corte"]) * eficiencia_media_maquinas_corte
    )

    return tempo_enfesto_ajustado + tempo_corte_ajustado


def gerar_cronograma(
    pedidos_ordenados, resultados, grades, data_inicio, recursos, turnos
):
    cronograma = {}
    tempo_atual_enfestamento = {r.id: data_inicio for r in recursos["enfestadeiras"]}
    tempo_atual_corte = {r.id: data_inicio for r in recursos["maquinas_corte"]}

    for p in pedidos_ordenados:
        resultado = resultados[p]

        # Enfestamento
        enfestadeira = min(tempo_atual_enfestamento, key=tempo_atual_enfestamento.get)
        inicio_enfestamento = tempo_atual_enfestamento[enfestadeira]
        fim_enfestamento = calcular_fim_operacao(
            inicio_enfestamento, resultado["tempo_enfesto"], turnos
        )
        tempo_atual_enfestamento[enfestadeira] = fim_enfestamento

        # Corte
        maquina_corte = min(tempo_atual_corte, key=tempo_atual_corte.get)
        inicio_corte = max(fim_enfestamento, tempo_atual_corte[maquina_corte])
        fim_corte = calcular_fim_operacao(
            inicio_corte, resultado["tempo_corte"], turnos
        )
        tempo_atual_corte[maquina_corte] = fim_corte

        cronograma[p] = {
            "inicio_enfestamento": inicio_enfestamento,
            "fim_enfestamento": fim_enfestamento,
            "enfestadeira": enfestadeira,
            "inicio_corte": inicio_corte,
            "fim_corte": fim_corte,
            "maquina_corte": maquina_corte,
        }

    return cronograma


def calcular_fim_operacao(inicio, duracao, turnos):
    tempo_atual = inicio
    tempo_restante = duracao

    while tempo_restante > 0:
        turno_atual = next(
            (t for t in turnos if t.inicio <= tempo_atual.time() < t.fim), None
        )
        if turno_atual:
            tempo_turno = min(
                (turno_atual.fim.hour - tempo_atual.time().hour) * 3600,
                tempo_restante * 3600,
            )
            tempo_restante -= tempo_turno / 3600 / turno_atual.eficiencia
            tempo_atual += datetime.timedelta(seconds=tempo_turno)
        else:
            tempo_atual += datetime.timedelta(days=1)
            tempo_atual = tempo_atual.replace(
                hour=turnos[0].inicio.hour, minute=0, second=0
            )

    return tempo_atual


def ajustar_grade(grade, tolerancia_largura):
    largura_total = sum(
        grade["larguras"][tamanho] * quantidade
        for tamanho, quantidade in grade["quantidades"].items()
    )
    if largura_total > grade["largura_tecido"] * (1 + tolerancia_largura):
        fator_ajuste = grade["largura_tecido"] / largura_total
        for tamanho in grade["quantidades"]:
            grade["quantidades"][tamanho] = max(
                1, int(grade["quantidades"][tamanho] * fator_ajuste)
            )
    return grade


def calcular_desperdicio(grade):
    largura_utilizada = sum(
        grade["quantidades"][t] * grade["larguras"][t] for t in grade["quantidades"]
    )
    return max(0, grade["largura_tecido"] - largura_utilizada)


def calcular_custo_por_camada(grade, area_total, perimetro_total, comprimento_enfesto):
    aproveitamento = grade["aproveitamento"]

    C_tecido = (area_total / aproveitamento) * grade["custo_tecido"]
    C_corte = perimetro_total * grade["custo_corte"]
    C_enfesto = (
        grade["custo_enfesto_fixo"]
        + comprimento_enfesto * grade["custo_enfesto_variavel"]
    )

    return C_tecido + C_corte + C_enfesto


def gerar_demanda_flutuante(demanda_base, variacao=0.05):
    random.seed(10)  # Reproducibilidade
    return int(demanda_base * (1 + random.uniform(-variacao, variacao)))


def otimizar_pedido(
    pedido,
    grades,
    tamanhos,
    comprimento_mesa_enfesto,
    recursos,
    percentual_superproducao,
    max_camadas_por_grade,
    fator_relaxacao,
    horas_producao,
    penalizacao_superproducao,
):
    # Cria um solver usando o SCIP
    solver = pywraplp.Solver.CreateSolver("SCIP")
    solver.SetTimeLimit(
        600_000
    )  # Define um limite de tempo de 10 minutos para a solução

    # Variáveis de decisão
    use_grade = {
        g: solver.BoolVar(f"use_{g}") for g in grades
    }  # Variável booleana para indicar se uma grade é usada
    x = {
        g: solver.IntVar(0, max_camadas_por_grade, f"x_{g}") for g in grades
    }  # Variável inteira para o número de camadas de cada grade
    superproducao = {
        t: solver.NumVar(0, solver.infinity(), f"super_{t}") for t in tamanhos
    }  # Variável para controlar a superprodução

    # Variáveis para recursos
    enfestadeiras = {
        r.id: solver.BoolVar(f"enfestadeira_{r.id}") for r in recursos["enfestadeiras"]
    }  # Variáveis para enfestadeiras
    maquinas_corte = {
        r.id: solver.BoolVar(f"maquina_corte_{r.id}")
        for r in recursos["maquinas_corte"]
    }  # Variáveis para máquinas de corte

    # Restrições para o uso das grades
    for g in grades:
        solver.Add(
            x[g] <= max_camadas_por_grade * use_grade[g]
        )  # Limita o número de camadas pela grade utilizada

    M_numero_de_grades = 10000  # Um número grande para garantir que a restrição não seja muito restritiva
    for g in grades:
        solver.Add(
            x[g] <= M_numero_de_grades * use_grade[g]
        )  # Restrições adicionais para o uso das grades

    # Restrições de demanda
    for t in tamanhos:
        solver.Add(
            solver.Sum(grades[g]["quantidades"][t] * x[g] for g in grades)
            >= pedido["demandas"][t]
        )  # Atende a demanda mínima
        solver.Add(
            solver.Sum(grades[g]["quantidades"][t] * x[g] for g in grades)
            <= pedido["demandas"][t] * (1 + percentual_superproducao) + superproducao[t]
        )  # Limita a superprodução

    # Restrições de comprimento do enfesto
    solver.Add(
        solver.Sum(grades[g]["comprimento_enfesto"] * x[g] for g in grades)
        <= comprimento_mesa_enfesto * fator_relaxacao * max_camadas_por_grade
    )

    # Cálculo do tempo de enfesto e corte
    tempo_enfesto = solver.Sum(
        grades[g]["tempo_enfesto_por_metro"] * grades[g]["comprimento_enfesto"] * x[g]
        for g in grades
    )
    tempo_corte = solver.Sum(
        grades[g]["tempo_corte_por_metro"] * grades[g]["perimetro_total"] * x[g]
        for g in grades
    )

    # Capacidade total de enfestamento e corte
    capacidade_enfestamento = solver.Sum(
        enfestadeiras[r.id] * r.eficiencia * horas_producao
        for r in recursos["enfestadeiras"]
    )
    capacidade_corte = solver.Sum(
        maquinas_corte[r.id] * r.eficiencia * horas_producao
        for r in recursos["maquinas_corte"]
    )

    # Restrições de capacidade
    solver.Add(
        tempo_enfesto <= capacidade_enfestamento
    )  # Restrições para o tempo de enfesto
    solver.Add(tempo_corte <= capacidade_corte)  # Restrições para o tempo de corte

    # Cálculo dos custos
    custo_producao = solver.Sum(
        grades[g]["custo_por_camada"] * x[g] for g in grades
    )  # Custo total de produção
    custo_setup_total = solver.Sum(
        grades[g]["custo_setup"] * use_grade[g] for g in grades
    )  # Custo total de setup
    custo_superproducao = solver.Sum(
        superproducao[t] * penalizacao_superproducao for t in tamanhos
    )  # Custo da superprodução

    # Função objetivo: minimizar o custo total
    objetivo = custo_producao + custo_setup_total + custo_superproducao
    solver.Minimize(objetivo)

    # Resolve o problema
    status = solver.Solve()

    # Verifica se uma solução viável foi encontrada
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        # Cálculo dos tempos reais de enfesto e corte
        tempo_enfesto_real = sum(
            grades[g]["tempo_enfesto_por_camada"] * x[g].solution_value()
            for g in grades
        )
        tempo_corte_real = sum(
            grades[g]["tempo_corte_por_camada"] * x[g].solution_value() for g in grades
        )
        tempo_total_real = tempo_enfesto_real + tempo_corte_real

        # Monta o resultado com as informações relevantes
        resultado = {
            "custo_total": solver.Objective().Value(),
            "custo_producao": custo_producao.solution_value(),
            "custo_setup": custo_setup_total.solution_value(),
            "camadas": {g: x[g].solution_value() for g in grades},
            "grades_usadas": [g for g in grades if use_grade[g].solution_value() > 0.5],
            "producao": {
                t: sum(
                    grades[g]["quantidades"][t] * x[g].solution_value() for g in grades
                )
                for t in tamanhos
            },
            "metros_tecido": sum(
                grades[g]["comprimento_enfesto"] * x[g].solution_value() for g in grades
            ),
            "perimetro_cortado": sum(
                grades[g]["perimetro_total"] * x[g].solution_value() for g in grades
            ),
            "superproducao": {t: superproducao[t].solution_value() for t in tamanhos},
            "desperdicio": sum(
                calcular_desperdicio(grades[g]) * x[g].solution_value() for g in grades
            ),
            "enfestadeiras_usadas": [
                r.id
                for r in recursos["enfestadeiras"]
                if enfestadeiras[r.id].solution_value() > 0.5
            ],
            "maquinas_corte_usadas": [
                r.id
                for r in recursos["maquinas_corte"]
                if maquinas_corte[r.id].solution_value() > 0.5
            ],
            "tempo_enfesto": tempo_enfesto_real,
            "tempo_corte": tempo_corte_real,
            "tempo_total": tempo_total_real,
            "fator_relaxacao": fator_relaxacao,
            "comprimento_enfesto_maximo": comprimento_mesa_enfesto
            * fator_relaxacao
            * max_camadas_por_grade,
        }
        return resultado
    else:
        logging.warning(
            f"Não foi possível encontrar uma solução ótima. Status: {status}"
        )
        return None


def ler_recursos():
    enfestadeiras = []
    maquinas_de_corte = []

    num_enfestadeiras = int(input("Número de enfestadeiras: "))
    for i in range(num_enfestadeiras):
        eficiencia = float(input(f"Eficiência da enfestadeira {i+1}: "))
        enfestadeiras.append(Recurso(id=f"E{i+1}", eficiencia=eficiencia))

    num_maquinas_corte = int(input("Número de máquinas de corte: "))
    for i in range(num_maquinas_corte):
        eficiencia = float(input(f"Eficiência da máquina de corte {i+1}: "))
        maquinas_de_corte.append(Recurso(id=f"C{i+1}", eficiencia=eficiencia))

    return enfestadeiras, maquinas_de_corte


def exportar_demanda_pedidos_excel(pedidos, nome_arquivo="demanda_pedidos.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Demanda dos Pedidos"

    # Cabeçalhos
    headers = ["ID do Pedido", "Prazo", "P", "M", "G", "GG"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

    # Dados dos pedidos
    for row, (pedido_id, pedido) in enumerate(pedidos.items(), start=2):
        ws.cell(row=row, column=1, value=pedido_id)
        ws.cell(row=row, column=2, value=pedido["prazo"])
        for col, tamanho in enumerate(["P", "M", "G", "GG"], start=3):
            ws.cell(row=row, column=col, value=pedido["demandas"][tamanho])

    # Ajustar largura das colunas
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Salvar o arquivo
    wb.save(nome_arquivo)
    print(f"Demanda dos pedidos exportada para {nome_arquivo}")


# Simula um cenário de produção, onde diferentes pedidos com quantidades variadas e prazos distintos precisam ser gerenciados e otimizados.
# TODO: Cada pedido vai estar associado a um conjunto de grades que por sua vez estará associado a aproveitamentos do tecido e este tecido terá
# seus custos de metro linear, por camada, por enfesto, por corte, etc. Além disso, os pedidos deverão ser passados em formatos json com os pazos e a data atual do planejamento
# deverá ser passada como parâmetro que fará parte do horizonte de planejamento.
def gerar_pedidos_para_intervalo(data_inicio, num_dias, pedidos_reais=None):
    if pedidos_reais:
        return pedidos_reais

    random.seed(10)  # Reproducibilidade
    pedidos = {}
    pedido_id = 1
    for dia in range(num_dias):
        data_atual = data_inicio + datetime.timedelta(days=dia)
        days1 = random.randint(1, 7)
        days2 = random.randint(1, 7)
        days3 = random.randint(1, 7)
        novos_pedidos = {
            pedido_id: {
                "demandas": {
                    t: gerar_demanda_flutuante(d, variacao=0.05)
                    for t, d in {"P": 50, "M": 100, "G": 80, "GG": 70}.items()
                },
                "prazo": data_atual + datetime.timedelta(days=days1),
            },
            pedido_id
            + 1: {
                "demandas": {
                    t: gerar_demanda_flutuante(d, variacao=0.05)
                    for t, d in {"P": 30, "M": 60, "G": 90, "GG": 40}.items()
                },
                "prazo": data_atual + datetime.timedelta(days=days2),
            },
            pedido_id
            + 2: {
                "demandas": {
                    t: gerar_demanda_flutuante(d, variacao=0.05)
                    for t, d in {"P": 40, "M": 100, "G": 100, "GG": 80}.items()
                },
                "prazo": data_atual + datetime.timedelta(days=days3),
            },
        }
        pedidos.update(novos_pedidos)
        pedido_id += 3
    return pedidos


def exportar_grades_excel(grades, nome_arquivo="grades_disponiveis.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades Disponíveis"

    # Cabeçalhos
    headers = [
        "Grade",
        "P",
        "M",
        "G",
        "GG",
        "Aproveitamento",
        "Custo Setup",
        "Área Total",
        "Comprimento Enfesto",
        "Perímetro Total",
        "Custo por Camada",
        "Desperdício",
        "Tempo Produção",
        "Tempo Corte",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

    # Dados das grades
    for row, (grade_nome, grade_info) in enumerate(grades.items(), start=2):
        ws.cell(row=row, column=1, value=grade_nome)
        for col, tamanho in enumerate(["P", "M", "G", "GG"], start=2):
            ws.cell(row=row, column=col, value=grade_info["quantidades"][tamanho])
        ws.cell(row=row, column=6, value=grade_info["aproveitamento"])
        ws.cell(row=row, column=7, value=grade_info["custo_setup"])
        ws.cell(row=row, column=8, value=grade_info.get("area_total", ""))
        ws.cell(row=row, column=9, value=grade_info.get("comprimento_enfesto", ""))
        ws.cell(row=row, column=10, value=grade_info.get("perimetro_total", ""))
        ws.cell(row=row, column=11, value=grade_info.get("custo_por_camada", ""))
        ws.cell(row=row, column=12, value=grade_info.get("desperdicio", ""))
        ws.cell(row=row, column=13, value=grade_info.get("tempo_producao", ""))
        ws.cell(row=row, column=14, value=grade_info.get("tempo_corte", ""))

    # Ajustar largura das colunas
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Salvar o arquivo
    wb.save(nome_arquivo)
    print(f"Grades disponíveis exportadas para {nome_arquivo}")


def exportar_informacoes_producao(
    custos, tamanhos, larguras, areas, nome_arquivo="informacoes_producao.xlsx"
):
    # Criar um novo workbook e uma nova planilha
    wb = Workbook()
    ws = wb.active
    ws.title = "Informações de Produção"

    # Cabeçalhos
    headers = ["Descrição", "Valor (R$)"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )

    # Adicionar custos à planilha
    for row, (descricao, valor) in enumerate(custos.items(), start=2):
        ws.cell(row=row, column=1, value=descricao)
        ws.cell(row=row, column=2, value=valor)

    # Adicionar informações sobre tamanhos, larguras e áreas
    row += len(custos) + 2  # Pular uma linha após os custos
    ws.cell(row=row, column=1, value="Tamanhos").font = Font(bold=True)
    ws.cell(row=row + 1, column=1, value="Tamanho")
    ws.cell(row=row + 1, column=2, value="Largura (m)")
    ws.cell(row=row + 1, column=3, value="Área (m²)")

    for i, tamanho in enumerate(tamanhos, start=2):
        ws.cell(row=row + i, column=1, value=tamanho)
        ws.cell(row=row + i, column=2, value=larguras[tamanho])
        ws.cell(row=row + i, column=3, value=areas[tamanho])

    # Ajustar largura das colunas
    for col in range(1, 4):
        ws.column_dimensions[get_column_letter(col)].auto_size = True

    # Salvar o arquivo
    wb.save(nome_arquivo)
    print(f"Informações de produção exportadas para {nome_arquivo}")


def main(
    criterio_prioridade,
    data_inicio,
    num_dias,
    recursos,
    tolerancia_largura,
    percentual_superproducao,
    max_camadas_por_grade,
    grades,
    pedidos,
    turnos,
    horas_producao,
    comprimento_mesa_enfesto,
    penalizacao_superproducao,
    relaxacao,
    pedidos_reais=None,
):

    diretorio_atual = os.getcwd()
    logging.info(f"Diretório atual: {diretorio_atual}")

    recursos_obj = {
        "enfestadeiras": [
            Recurso(r["id"], r["eficiencia"]) for r in recursos["enfestadeiras"]
        ],
        "maquinas_corte": [
            Recurso(r["id"], r["eficiencia"]) for r in recursos["maquinas_corte"]
        ],
    }

    # Utilizado somente para a função  exportar_informacoes_producao
    larguras = {"P": 0.2, "M": 0.22, "G": 0.23, "GG": 0.24}  # em metros
    areas = {
        "P": 0.2 * 0.7,
        "M": 0.22 * 0.8,
        "G": 0.23 * 0.9,
        "GG": 0.24 * 1.0,
    }  # em m² (aprox.)

    comprimento_mesa_enfesto = (
        comprimento_mesa_enfesto  #  Parâmetro de entrada para a Otimização
    )

    # Utilizado somente para a função  exportar_informacoes_producao
    custos = {
        "Custo do tecido por metro linear": 18.90,
        "Custo de corte por metro de perímetro": 0.45,
        "Custo fixo de enfestamento": 1.25,
        "Custo variável de enfestamento por metro de enfesto": 0.30,
    }

    tamanhos = list(grades["Grade1"]["quantidades"].keys())
    exportar_informacoes_producao(
        custos, tamanhos, larguras, areas, "informacoes_producao.xlsx"
    )

    # Exportar grades disponíveis para Excel
    exportar_grades_excel(grades, "grades_disponiveis.xlsx")

    for g in grades:
        grades[g] = ajustar_grade(grades[g], tolerancia_largura)
        quantidades = grades[g]["quantidades"]
        aproveitamento = grades[g]["aproveitamento"]

        area_total = sum(quantidades[s] * grades[g]["areas"][s] for s in tamanhos)
        L = area_total / (grades[g]["largura_tecido"] * aproveitamento)
        P_total = sum(quantidades[s] * grades[g]["perimetros"][s] for s in tamanhos)

        tempo_enfesto = L * grades[g]["tempo_enfesto_por_metro"]
        tempo_corte = P_total * grades[g]["tempo_corte_por_metro"]
        tempo_total_por_camada = tempo_enfesto + tempo_corte
        grades[g]["tempo_producao"] = tempo_total_por_camada
        tempo_por_camada = tempo_enfesto + tempo_corte
        max_camadas = grades[g].get("max_camadas", 1)
        tempo_total = tempo_por_camada * max_camadas

        grades[g].update(
            {
                "area_total": area_total,
                "comprimento_enfesto": L,
                "perimetro_total": P_total,
                "custo_por_camada": calcular_custo_por_camada(
                    grades[g], area_total, P_total, L
                ),
                "desperdicio": calcular_desperdicio(grades[g]),
                "tempo_enfesto_por_camada": tempo_enfesto,
                "tempo_corte_por_camada": tempo_corte,
                "tempo_por_camada": tempo_por_camada,
                "tempo_total_maximo": tempo_total,
                "max_camadas": max_camadas,
            }
        )

    logging.info(
        f"Grade {g}: Área total = {area_total:.2f} m², Comprimento do enfesto = {L:.2f} m, "
        f"Tempo por camada = {tempo_por_camada:.2f} h, Tempo total máximo = {tempo_total:.2f} h"
    )

    # pedidos = gerar_pedidos_para_intervalo(data_inicio, num_dias, pedidos_reais)

    resultados = {}
    for p, pedido in pedidos.items():
        logging.info(f"Otimizando pedido {p}")
        logging.info(f"Demandas: {pedido['demandas']}")
        # resultado = otimizar_pedido_com_relaxacao(pedido, grades, tamanhos, comprimento_maximo_enfesto, recursos_obj, percentual_superproducao=0.05, max_camadas_por_grade=30)
        resultado = otimizar_pedido_com_relaxacao(
            pedido,
            grades,
            tamanhos,
            comprimento_mesa_enfesto,
            recursos_obj,
            percentual_superproducao=percentual_superproducao,
            max_camadas_por_grade=max_camadas_por_grade,
            horas_producao=horas_producao,
            penalizacao_superproducao=penalizacao_superproducao,
            relaxacao=relaxacao,
        )

        if resultado:
            # Verificar se a produção atende à demanda
            demanda_atendida = all(
                resultado["producao"][t] >= pedido["demandas"][t] for t in tamanhos
            )
            if demanda_atendida:
                resultados[p] = resultado
                logging.info(
                    f"Solução encontrada para o pedido {p} com relaxação de {resultado['relaxacao_aplicada']:.2f}"
                )
            else:
                logging.warning(
                    f"Solução encontrada para o pedido {p}, mas não atende completamente à demanda. Relaxação aplicada: {resultado['relaxacao_aplicada']:.2f}"
                )
        else:
            logging.warning(
                f"Não foi possível encontrar uma solução viável para o Pedido {p}, mesmo com relaxação"
            )
    
        

    def calcular_prioridade(pedido, resultado, criterio):
        if criterio == "prazo":
            return (pedido["prazo"] - data_inicio).days
        elif criterio == "custo_total":
            return resultado["custo_total"]
        elif criterio == "tempo_producao":
            return calcular_tempo_producao(resultado, grades, recursos_obj)
        else:
            raise ValueError(f"Critério de prioridade inválido: {criterio}")

    prioridades = {
        p: calcular_prioridade(pedidos[p], resultado, criterio_prioridade)
        for p, resultado in resultados.items()
    }

    pedidos_ordenados = sorted(prioridades, key=prioridades.get)

    cronograma = gerar_cronograma(
        pedidos_ordenados, resultados, grades, data_inicio, recursos_obj, turnos
    )

    # Exibir resultados e gráfico
    print(f"\nPriorização dos pedidos ({criterio_prioridade}):")
    for i, p in enumerate(pedidos_ordenados, 1):
        print(f"{i}. Pedido {p}: {prioridades[p]:.2f}")

    print("\nCronograma de produção:")
    for p in pedidos_ordenados:
        print(f"Pedido {p}:")
        print(
            f"  Enfestamento: Início {cronograma[p]['inicio_enfestamento'].strftime('%d/%m/%Y %H:%M')} - "
            f"Fim {cronograma[p]['fim_enfestamento'].strftime('%d/%m/%Y %H:%M')} "
            f"(Enfestadeira {cronograma[p]['enfestadeira']})"
        )
        print(
            f"  Corte: Início {cronograma[p]['inicio_corte'].strftime('%d/%m/%Y %H:%M')} - "
            f"Fim {cronograma[p]['fim_corte'].strftime('%d/%m/%Y %H:%M')} "
            f"(Máquina de Corte {cronograma[p]['maquina_corte']})"
        )

    print("\nDetalhes dos pedidos:")
    for p in pedidos_ordenados:
        resultado = resultados[p]
        print(f"\nPedido {p}:")
        print(f"  Prazo: {pedidos[p]['prazo'].strftime('%d/%m/%Y')}")
        print(f"  Custo total: R$ {resultado['custo_total']:.2f}")
        print(f"  Metros de tecido: {resultado['metros_tecido']:.2f} m")
        print(f"  Perímetro cortado: {resultado['perimetro_cortado']:.2f} m")
        print(f"  Desperdício de tecido: {resultado['desperdicio']:.2f} m²")
        print(f"  Tempo de enfesto real: {resultado['tempo_enfesto']:.2f} h")
        print(f"  Tempo de corte real: {resultado['tempo_corte']:.2f} h")
        print(f"  Tempo total real: {resultado['tempo_total']:.2f} h")
        print("  Produção:")
        for t in tamanhos:
            print(
                f"    {t}: {int(resultado['producao'][t])} peças (Demanda: {pedidos[p]['demandas'][t]})"
            )
        print("  Camadas utilizadas:")
        for g, camadas in resultado["camadas"].items():
            if camadas > 0:
                print(f"    {g}: {int(camadas)} camadas")

        print(f"  Custo de setup: R$ {resultado['custo_setup']:.2f}")

    fig = criar_grafico_gantt(cronograma, pedidos_ordenados, criterio_prioridade)
    pio.show(fig)

    exportar_para_excel(
        cronograma, pedidos_ordenados, resultados, pedidos, tamanhos, grades, fig
    )
    # Exportar a demanda dos pedidos para Excel
    exportar_demanda_pedidos_excel(
        pedidos, f"demanda_pedidos-{criterio_prioridade}.xlsx"
    )
    salvar_figura_html(fig, f"gantt_chart-{criterio_prioridade}.html")


    # Preparar resultados detalhados
    resultados_detalhados = {}
    for p in pedidos_ordenados:
        resultado = resultados[p]
        resultados_detalhados[p] = {
            "prazo": pedidos[p]["prazo"].isoformat(),
            "custo_total": float(resultado["custo_total"]),
            "metros_tecido": float(resultado["metros_tecido"]),
            "perimetro_cortado": float(resultado["perimetro_cortado"]),
            "desperdicio": float(resultado["desperdicio"]),
            "tempo_enfesto": float(resultado["tempo_enfesto"]),
            "tempo_corte": float(resultado["tempo_corte"]),
            "tempo_total": float(resultado["tempo_total"]),
            "producao": {t: int(resultado["producao"][t]) for t in tamanhos},
            "demandas": pedidos[p]["demandas"],
            "camadas": {g: int(camadas) for g, camadas in resultado["camadas"].items() if camadas > 0},
            "custo_setup": float(resultado["custo_setup"])
        }

    # Retornar resultados
    return {
        "pedidos_ordenados": pedidos_ordenados,
        "cronograma": {
            str(k): {
                "inicio_enfestamento": v["inicio_enfestamento"].isoformat(),
                "fim_enfestamento": v["fim_enfestamento"].isoformat(),
                "enfestadeira": v["enfestadeira"],
                "inicio_corte": v["inicio_corte"].isoformat(),
                "fim_corte": v["fim_corte"].isoformat(),
                "maquina_corte": v["maquina_corte"]
            }
            for k, v in cronograma.items()
        },
        "resultados": resultados_detalhados,
        "metricas_globais": {
            "custo_total": sum(r["custo_total"] for r in resultados_detalhados.values()),
            "tempo_total": sum(r["tempo_total"] for r in resultados_detalhados.values()),
            "desperdicio_total": sum(r["desperdicio"] for r in resultados_detalhados.values())
        }
    }


if __name__ == "__main__":
    main()